#!/usr/bin/env python3
"""
roc_results.py — adapter for the "ROC Results" enrichment sheet.

This is the layer that was missing. A raw ROC export has no contact data, so
nothing in it can be built. The ROC Results tab is that export after a lookup
pass, and it carries phone on ~99.9% of records.

It is NOT a flat table. Two row shapes share a License Number column:

  primary    License No | Business Name | Officers | Class | QP | Status | Address | Phone | TS
  secondary  License No | "ROC 322306"  | Class    | Status | "More Info" | ...

A secondary row is an ADDITIONAL license held by the same business, and the
columns mean different things than the header says. Reading the sheet flat
produces phantom contractors named "ROC 322306" and silently drops the trade
signal those rows carry.

That signal matters: 98 businesses whose primary class is general (B / KB-*)
hold a specialty license that reveals what they actually do. Stark Contrast LLC
is KB-2 on paper and CR-21 Hardscaping in practice.

    python3 roc_results.py results.xlsx --sheet "ROC Results" --out ./out
"""

import argparse
import collections
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from acp_schema import ACP, digits, format_phone, slugify, split_list  # noqa: E402

EXPECTED_HEADERS = ["License Number", "Business Name", "Name and Title", "Class",
                    "Qualifying Party", "Status", "Address", "Phone", "Timestamp"]

GENERAL_CODE = re.compile(r"^(A|B|KA|KB)(-\d+)?$")
SUB_NAME = re.compile(r"^ROC\s*\d+$", re.I)
PUBLISHABLE_STATUS = {"active"}


def is_secondary(row) -> bool:
    """Secondary rows carry 'More Info' in col 5 and 'ROC <number>' as the name."""
    return (str(row[4]).strip() == "More Info") or bool(SUB_NAME.match(str(row[1]).strip()))


def split_class(value):
    """'CR-77 Plumbing Including Solar' -> ('CR-77', 'Plumbing Including Solar')"""
    v = str(value or "").strip()
    if not v:
        return "", ""
    parts = v.split(" ", 1)
    code = parts[0].strip().upper()
    return code, (parts[1].strip() if len(parts) > 1 else "")


def split_dba(value):
    """'Any Hour LLCDBA : AC by J' -> ('Any Hour LLC', 'AC by J')"""
    v = str(value or "").strip()
    m = re.split(r"\s*DBA\s*:\s*", v, maxsplit=1, flags=re.I)
    if len(m) == 2:
        return m[0].strip(), m[1].strip()
    return v, ""


def parse_address(value):
    """'Tucson, AZ, 85713' -> ('Tucson','AZ','85713'). Tolerates empty segments."""
    parts = [p.strip() for p in str(value or "").split(",")]
    parts = [p for p in parts if p]          # 'Miami Lakes,, FL, 33016'
    city = state = zip_ = ""
    if parts:
        zip_ = digits(parts[-1])[:5] if digits(parts[-1]) else ""
        rest = parts[:-1] if zip_ else parts
        if rest:
            last = rest[-1].upper()
            if len(last) == 2 and last.isalpha():
                state, rest = last, rest[:-1]
        city = ", ".join(rest)
    return city, state, zip_


def parse_officers(value):
    """
    'Ann Diaz (Member) Ann Diaz (Qualifying Party) Bo Lee (Owner)'
      -> [('Ann Diaz', ['Member','Qualifying Party']), ('Bo Lee', ['Owner'])]
    Order is preserved so the first human named stays first.
    """
    text = str(value or "").strip()
    if not text:
        return []
    people, order = collections.OrderedDict(), []
    for m in re.finditer(r"([^()]+?)\s*\(([^)]+)\)", text):
        name = re.sub(r"\s+", " ", m.group(1)).strip(" ,;")
        role = m.group(2).strip()
        if not name:
            continue
        if name not in people:
            people[name] = []
            order.append(name)
        if role not in people[name]:
            people[name].append(role)
    return [(n, people[n]) for n in order]


def pick_owner(officers, qualifying_party=""):
    """Prefer the qualifying party; fall back to the first human-looking name."""
    qp = str(qualifying_party or "").strip()
    if qp:
        return qp
    for name, roles in officers:
        if "Parent Entity" in roles:
            continue
        if re.search(r"\b(LLC|INC|CORP|COMPANY|LP|LTD)\b", name, re.I):
            continue
        return name
    return ""


def read_sheet(path: Path, sheet_name: str):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        sys.exit(f"sheet '{sheet_name}' not found. sheets: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = []
    for r in ws.iter_rows(values_only=True):
        vals = ["" if v is None else str(v).strip() for v in r]
        vals += [""] * (9 - len(vals))
        if any(vals):
            rows.append(vals[:9])
    header, body = rows[0], rows[1:]
    missing = [h for h in EXPECTED_HEADERS if h not in header]
    if missing:
        print(f"  ! header drift, missing: {missing}", file=sys.stderr)
    return body


def group_rows(rows):
    """
    Group by License Number, splitting primary from secondary.
    Returns [{primary, secondaries}] in first-seen order.
    """
    groups = collections.OrderedDict()
    for row in rows:
        lic = str(row[0]).strip()
        if not lic:
            continue
        g = groups.setdefault(lic, {"primary": None, "secondaries": []})
        if is_secondary(row):
            g["secondaries"].append({
                "roc_number": digits(row[1]),
                "class_code": split_class(row[2])[0],
                "class_detail": split_class(row[2])[1],
                "status": str(row[3]).strip(),
            })
        elif g["primary"] is None:
            g["primary"] = row
        # Duplicate primaries (23 in the Dec-2025 file) are the same business
        # listed twice; the first wins rather than creating a second client.
    return [g for g in groups.values() if g["primary"]]


def to_client(acp: ACP, group, tier="lite"):
    row = group["primary"]
    secondaries = group["secondaries"]

    business, dba = split_dba(row[1])
    officers = parse_officers(row[2])
    code, detail = split_class(row[3])
    city, state, zip_ = parse_address(row[6])
    owner = pick_owner(officers, row[4])

    c = acp.blank()
    c["tier"] = tier
    c["business_name"] = business
    c["short_name"] = dba or ""
    c["owner"] = owner
    c["owner_first_name"] = owner.split(" ")[0] if owner else ""
    c["roc_number"] = digits(row[0])
    c["license_class"] = code
    c["license_class_description"] = detail
    c["roc_status"] = str(row[5]).strip() or "Active"
    c["phone"] = format_phone(row[7])
    c["city"] = city
    c["state"] = state or "AZ"
    c["zip"] = zip_
    c["service_area"] = [city] if city else []

    # Only currently-held licenses count as a trade signal.
    also = [s["class_code"] for s in secondaries
            if s["status"].strip().lower() in PUBLISHABLE_STATUS and s["class_code"]]
    trade, conf, reason = acp.infer_trade(code, detail, business, "", also_holds=also)
    c["trade"] = trade
    c["trade_confidence"] = conf

    c["roc"] = {
        "primary_class": code,
        "primary_class_detail": detail,
        "also_holds": [
            {"roc_number": s["roc_number"], "class": s["class_code"],
             "detail": s["class_detail"], "status": s["status"]}
            for s in secondaries
        ],
        "officers": [{"name": n, "roles": r} for n, r in officers],
        "dba": dba,
        "trade_reason": reason,
    }
    return c


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--sheet", default="ROC Results")
    ap.add_argument("--out", default="./roc-results-out")
    ap.add_argument("--metro", help="comma-separated city allow-list")
    ap.add_argument("--tier", choices=["lite", "full"], default="lite")
    args = ap.parse_args()

    acp = ACP()
    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)

    rows = read_sheet(Path(args.workbook), args.sheet)
    groups = group_rows(rows)
    clients = [acp.resolve(to_client(acp, g, args.tier)) for g in groups]

    if args.metro:
        allow = {c.strip().lower() for c in args.metro.split(",")}
        clients = [c for c in clients if c["city"].strip().lower() in allow]

    ready, blocked, review = [], [], []
    for c in clients:
        state, missing = acp.readiness(c)
        c["_readiness"] = state
        c["_missing"] = missing
        if c["roc_status"].strip().lower() not in PUBLISHABLE_STATUS:
            blocked.append(c)
        elif c["trade_confidence"] == "low":
            review.append(c)
        elif state == "prospect":
            blocked.append(c)
        else:
            ready.append(c)

    (out / "ready.json").write_text(ACP.dumps(ready), encoding="utf-8")
    (out / "review.json").write_text(ACP.dumps(review), encoding="utf-8")
    (out / "blocked.json").write_text(ACP.dumps(blocked), encoding="utf-8")

    sharpened = [c for c in clients
                 if "sharpened by" in c.get("roc", {}).get("trade_reason", "")]
    by_trade = collections.Counter(c["trade"] for c in clients)
    lines = [
        f"sheet rows            {len(rows)}",
        f"businesses            {len(groups)}",
        f"after metro filter    {len(clients)}",
        "",
        f"BUILDABLE NOW         {len(ready)}",
        f"needs trade review    {len(review)}",
        f"blocked               {len(blocked)}",
        f"  non-active license  {sum(1 for c in blocked if c['roc_status'].strip().lower() not in PUBLISHABLE_STATUS)}",
        f"  missing contact     {sum(1 for c in blocked if c['_readiness'] == 'prospect')}",
        "",
        f"trade sharpened by secondary license   {len(sharpened)}",
        "",
        "by trade:",
        *[f"  {t:16} {n}" for t, n in by_trade.most_common()],
    ]
    summary = "\n".join(lines)
    (out / "summary.txt").write_text(summary + "\n", encoding="utf-8")
    print(summary)


if __name__ == "__main__":
    main()
